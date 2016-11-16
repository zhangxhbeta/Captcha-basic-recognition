# -*- coding:utf-8 -*-

from openpyxl import Workbook
import os

try:
    import xml.etree.cElementTree as et
except ImportError:
    import xml.etree.ElementTree as et

def process_all(dir, sheet):
    last_year = None
    row = 1
    for file in os.listdir(dir):
        if file.startswith('.'):
            continue

        p = file.split('.')[0].split('-')
        year = p[0]
        no = p[1]

        filepath = os.path.join(dir, file)

        if last_year is not None and last_year != year:
            sheet = wb.create_sheet(year)
            row = 1

        # 读取xml
        tree = et.ElementTree(file=filepath)
        root = tree.getroot()

        for elem in root.findall('row'):
            for index, cell in enumerate(elem.iter()):
                sheet.cell(row=row, column=index + 1, value=cell.text)
            row += 1

        last_year = year

if __name__ == "__main__":
    dest_filename = 'isisn_429.xlsx'

    years = ['1997']
    wb = Workbook()

    sheet = wb.active
    sheet.title = years[0]

    process_all('Isisn', sheet)

    wb.save(filename=dest_filename)



